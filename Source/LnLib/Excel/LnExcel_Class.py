#!/usr/bin/python3.4
# -*- coding: iso-8859-15 -*-
# -O Optimize e non scrive il __debug__
#
# ####################################################################################################################
import sys, os
import types
import platform
import openpyxl

from    openpyxl.utils import get_column_letter
import  openpyxl.utils as xls

from ..LnCommon.LnLogger import SetLogger        # OK funziona dalla upperDir del package

##########################################################################################################
# Per la creazione di attributi read-only
#   http://stackoverflow.com/questions/9920677/how-should-i-expose-read-only-fields-from-python-classes
#   http://stackoverflow.com/questions/14594120/python-read-only-property
##########################################################################################################

def ro_property(name):
    def ro_property_decorator(c):
        setattr(c, name, property(lambda o: o.__dict__["_" + name]))
        return c
    return ro_property_decorator

# creazione delle variabili che si voglio mettere in R/O
@ro_property('name')
@ro_property('address')
@ro_property('description')
@ro_property('author')



# ###########################################################
# - getInterfacesData(ifc) - con un solo indirizzo IP sulla scheda
# ###########################################################
class Excel(object):
        # ***********************************************
        # * Calling Sample:
        # * eth0 = Ln.LnInterfaces('eth0', myIP=True, DRYRUN=True, setLogger=gv.Ln.setLogger)
        # ***********************************************
    def __init__(self, excelFileName, fDEBUG=False):
            # ----- defaults
        self._name              = None
        self._description       = None
        self._author            = None
        self._setLogger         = None
        self._fDEBUG            = fDEBUG

        self._filename          = excelFileName
        self._description       = "class to manage execl file."
        self._author            = "Loreto Notarantonio"
        self._SetLogger         = SetLogger


        self._read()

    #######################################################
    # - read()
    # - Lettura di un file Excel e ritorna il WorkBook
    #######################################################
    def _read(self, keep_vba=False):
        logger = self._SetLogger(__name__)
        try:
                # warnings.simplefilter("ignore")
                # in read-only=True:
                #       colNames = ws.rows[1]
                #       TypeError: 'generator' object is not subscriptable
            self._wb = openpyxl.load_workbook(  self._filename,
                                                read_only=True,
                                                keep_vba=False,
                                                data_only=True
                                            )
                                                # use_iterators=False,
            self.sheetNames    = self._wb.get_sheet_names()

        except Exception as why:
            print("error reading file: {0} [{1}]".format(self._filename, why))
            logger.error("error reading file: {0} [{1}]".format(self._filename, why))
            sys.exit(99)


        if self._fDEBUG:
            print('sheet names: {0}'.format(self.sheetNames))
            logger.info('sheet names: {0}'.format(self.sheetNames))




    ####################################################################
    # - exportToCSV()
    # - Export di un foglio excel to CSV format
    # - rangeString: range di celle su cui operare
    # - colNames   : numero di riga che contiene i nomi delle colonne
    # - csvType     : simple    "a"; "b",...
    # -             : listtype  ["a"; "b",...]
    ####################################################################
    def exportCSV(self, sheetName, csvType='simple', outFname=None, rangeString=None, colNames=0, maxrows=99999999, fPRINT=False):
        logger = self._SetLogger(__name__)
        if fPRINT:
            print("Converting sheetName: [{0}] to CSV file: [{1}]." .format(sheetName, outFname))

        ws          = self._wb.get_sheet_by_name(sheetName)
        nRows       = ws.max_row
        nCols       = ws.max_column


        if rangeString:
            minCol, minRow, maxCol, maxRow = xls.range_boundaries(rangeString)
        else:
            minCol, minRow, maxCol, maxRow = 1, 1, ws.max_column, ws.max_row

        fullRange = get_column_letter(minCol) + str(minRow) + ':' + get_column_letter(maxCol) + str(maxRow)
        logger.info("     full Range: {0}".format(fullRange))

        minCol -= 1         # col parte da '0'
        maxCol -= 1         # col parte da '0'


            # ---------------------------------
            # - grosso modo può andare.....
            # ---------------------------------
        dataList        = []
        # dataListOfList  = []
        for indexRow, row in enumerate(ws.rows):
                # - prendiamo tutte le righe previste nel range
            if minRow <= indexRow < maxRow:
                # - ...e lo stesso per le colonne
                if indexRow >= colNames:
                    line = []
                    for indexCol, cell in enumerate(row):
                        if minCol <= indexCol <= maxCol:
                            val = cell.value if cell.value else ''
                            line.append(val)
                else:
                    continue

                if csvType == 'listtype':
                    dataList.append(line)

                else:
                        # costruiamo la riga ...
                    lineStr = line[0]
                    for item in line[1:]:
                        lineStr = '{0};{1}'.format(lineStr, item)

                        # ... per inserirla nell'array
                    dataList.append(lineStr)

        if fPRINT:
            for index, line in enumerate(dataList):
                print ('{0:5} - {1}'.format(index, line))

        if outFname:
            FILE = open(outFname, "wb")
            for line in dataList:
                line = "{0}{1}".format(line, '\n')
                FILE.write(bytes(line, 'UTF-8'))       # con Python3 bisogna convertirlo in bytes

            FILE.close()
            logger.info("..... file: {FILE} has been written".format(FILE=outFname))

            if fPRINT:
                print("..... file: {FILE} has been written".format(FILE=outFname))

        if fPRINT:
            print()
            print("     full Range: {0}".format(fullRange))
            print("     file {0} has been created".format(outFname))
            print()

            for item in dataList:
                print (item)
            print ()

        self.data = dataList











#######################################################
# -
#######################################################

if __name__ == "__main__":
    excelFileName = 'J:\\GIT-REPO\\Python3\\MP3Catalog\\data\\MP3_Master_forTEST.xlsm'
    csvFile = excelFileName.rsplit('.', -1)[0] + '.csv'
    mydata  = Excel(excelFileName)
    mydata.exportToCSV('Catalog', outFname=csvFile, rangeString="B2:Z17", colNames=4, fPRINT=True)
